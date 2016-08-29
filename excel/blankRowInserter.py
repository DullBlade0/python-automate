#! python
#blankRowInserter.py - takes three command line arguments, n for the row blank spaces start, m for the amount of blank rows, and an excel file.

import openpyxl, sys

n = int(sys.argv[1])
m = int(sys.argv[2])
wb1 = openpyxl.load_workbook(sys.argv[3])
wb2 = openpyxl.Workbook()

oldSheet = wb1.active 
newSheet = wb2.active 
oldList = []


#This loop is made to get the entire worksheet as a list

for list in oldSheet:
	miniList = []
	for cell in list:
		miniList.append(cell.value)
	oldList.append(miniList)

for i in range(0, m):
	oldList.insert(n, [])

#This loop is to add the blank spaces into the new file/worksheet
for x, items in enumerate(oldList, start=1):
	for y, item in enumerate(items, start=1):
		newSheet.cell(row = i, column = y).value = item
	i += 1

wb2.save('blankSpaced-'+ sys.argv[3])

