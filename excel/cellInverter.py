#! python3
# cellInverter.py - Inverts the cell of a spreadsheet

import openpyxl

wb1 = openpyxl.load_workbook('example.xlsx')
wb2 = openpyxl.Workbook()
sheet1 = wb1.active
sheet2 = wb2.active

for x in range(1, sheet1.max_row+1):
	for y in range(1, sheet1.max_column+1):
		sheet2.cell(row = x, column = y).value = sheet1.cell(row = y, column = x).value
		sheet2.cell(row = y, column = x).value = sheet1.cell(row = x, column = y).value

wb2.save('test.xlsx')